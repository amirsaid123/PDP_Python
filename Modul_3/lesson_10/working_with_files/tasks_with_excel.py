import openpyxl, csv

# path = "database/india.xlsx"
# wb_obj = openpyxl.load_workbook(path)
# sheet = wb_obj.active
#
# rows = sheet.max_row
# cols = sheet.max_column
#
#
# for i in range(1, rows+1):
#     for j in range(1, cols+1):
#         obj = sheet.cell(row=i, column=j).value
#         if j == 1 or j == 4:
#             print(obj, end=' ')
#     print()


data = []
with open("userss.csv", "r") as file:
    reader = csv.reader(file)
    for row in reader:
        data.append(row)

workbook = openpyxl.Workbook()
worksheet = workbook.active

for row_idx, row in enumerate(data):
    for col_idx, value in enumerate(row):
        worksheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

workbook.save("output.xlsx")