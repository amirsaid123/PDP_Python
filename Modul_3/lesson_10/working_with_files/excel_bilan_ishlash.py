import openpyxl

path = "database/workers.xlsx"
# wb = openpyxl.load_workbook(path)
# sheet_obj = wb.active
# row = sheet_obj.max_row
# col = sheet_obj.max_column
#

# for i in range(row):
#     for j in range(col):
#         obj = sheet_obj.cell(row=i+1, column=j+1)
#         print(obj.value, end=" ")
#     print()



wb = openpyxl.load_workbook("sample.xlsx")
sheet = wb.active
data = (
    (1, 2, 3),
    (4, 5, 6),
    (7, 8, 9),
)

for row in data:
    sheet.append(row)

wb.save("sample.xlsx")

#
# c1 = sheet.cell(row=1, column=1)
# c1.value = "Hello"
# c2 = sheet.cell(row=1, column=2)
# c2.value = "World"
# c3 = sheet["A2"]
# c3.value = "Welcome"
# c4 = sheet["B2"]
# c4.value = "Everyone"
# wb.save("sample.xlsx")